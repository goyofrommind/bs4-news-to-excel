from itertools import count
import requests
from bs4 import BeautifulSoup
import openpyxl

def gy_soup(url):  #res url ,headers 함수로 만들기
    headers = ({'User-Agent':'Mozilla/5.0'})
    res = requests.get(url, headers=headers)
    res.raise_for_status()
    soup = BeautifulSoup(res.text, "lxml")
    return soup

# word = input("검색할 명령어를 입력하세요 : ")
# print("-"*100)
word ='벚꽃'


try:
    wb = openpyxl.load_workbook("검색어_{}뉴스_.xlsx".format(word))
    sheet = wb.active 
    wb.create_sheet('검색어 뉴스')
    print("파일 갱신 완료")
except :
    wb = openpyxl.Workbook()
 
    sheet = wb.active
    sheet.append(["제목", "언론사",])
    sheet
   
    print("새로운 파일 생성 완료")


def word_news():
    for page in range(1, 100, 10):    # url 방식을 보면 페이지가 주소가 1/11/21 이런식으로 올라감.  따라서 for in range(start , stop , step)
        url = 'https://search.naver.com/search.naver?sm=tab_hty.top&where=news&query='+word+'&start='+str(page)
        soup = gy_soup(url)
        # print(url)
        # 음..첫번째 기사만 가져오는 문제를 발견..나머지 1~10 기사들 클래스제목이 다 동일함. 어케해야할까..?
        
        # news_check= soup.find_all('li', attrs={"class":"li"})
        # for j in news_check():
        news_title = soup.find("a", attrs = {"class":"news_tit"}).get_text()
        news_source = soup.find("a", attrs = {"class":"info press"}).get_text()

        print(news_title, news_source, sep='    -   ')
        sheet.append([news_title, news_source])


print("-" * 100)
if __name__ == "__main__":
    word_news()



wb.save("검색어_{}뉴스_.xlsx".format(word))
 

